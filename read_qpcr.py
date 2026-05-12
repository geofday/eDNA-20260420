import openpyxl

path = r'C:\Users\geofd\OneDrive\Documents\eDNA_2025\Union_River_2025_UMaine_GeoffdayqPCRresults.xlsx'
wb = openpyxl.load_workbook(path)
print("Sheets:", wb.sheetnames)

for shname in wb.sheetnames:
    ws = wb[shname]
    print(f"\n=== Sheet: {shname} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if any(v is not None for v in row):
            print(f"  row {i+1}: {row}")
        if i >= 30:
            print("  ... (truncated at row 31)")
            break
