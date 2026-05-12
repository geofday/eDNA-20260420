import openpyxl

path = r'C:\Users\geofd\OneDrive\Documents\eDNA_2025\Union_River_2025_UMaine_GeoffdayqPCRresults.xlsx'
wb = openpyxl.load_workbook(path)

for shname in wb.sheetnames:
    ws = wb[shname]
    print(f"\n=== {shname} — all rows ===")
    for row in ws.iter_rows(min_row=1, values_only=True):
        if any(v is not None for v in row):
            print(f"  {row}")
